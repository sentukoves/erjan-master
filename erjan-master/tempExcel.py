import openpyxl
import sys
import os
# import requests
# import json
from flask import Flask, render_template, redirect, json ,jsonify, request as re
import sqlite3
import datetime
import re as regular

app = Flask(__name__)
# ---------------------------------- Читаем excel и записываем в исходный массив -----------------------------------
# @staticmethod функция может использоваться в не класса
def parse_array(sql):
    path_db = os.path.join(sys.path[0], 'static/base/base.db')
    connect_db = sqlite3.connect(path_db)
    cursor = connect_db.cursor()
    print(sql)
    cursor.execute(sql)
    if sql.find("INSERT") > -1:
        connect_db.commit()
    if sql.find("SELECT") > -1:
        items = []
        for items_array in cursor.fetchall():
            items.append(items_array[0])
        return sorted(items)
class ReadExcel(object):
    def __init__(self):
        super(ReadExcel, self).__init__()

    @staticmethod
    def OpenFile():
        path_file = os.path.join(sys.path[0], "excel\асумтр.xlsx")
        read_file = openpyxl.load_workbook(path_file)
        sheetnames = read_file.sheetnames[0]
        read_file = read_file[sheetnames]
        max_row = read_file.max_row
        max_column = read_file.max_column
        array = []
        for i in range(max_row):
            array_temp = []
            for j in range(max_column):
                array_temp.append(str(read_file.cell(i + 1, j + 1).value))
            array.append(array_temp)
            del array_temp

        return array[1:]

class ParseJson(ReadExcel):
    def __init__(self):
        super(ParseJson, self).__init__()
        self.array = super().OpenFile()

    def make_array(self):
        slovar = {}

        for i, item in enumerate(self.array):
            slovar[item[0]] = []

        for key, Item in slovar.items():
            values = []
            for parse in self.array:
                if key == parse[0]:
                    values.append(parse[1:])
            slovar[key] = values
            del values

        return slovar

        #
        # request = requests.get(
        #     'https://speller.yandex.net/services/spellservice.json/checkText?text=={}'.format(values))
        # try:
        #     response_json = request.json()
        # except Exception.args:
        #     print(sys.path)
        # else:
        #     if response_json:
        #         parse_json = json.loads(response_json)
        #          print(parse_json)


# ----------------------------------------------------------------------------------------


# сверка с массиовом

class QeryStatsFileParse(object):

    def __init__(self, id=1):
        super(QeryStatsFileParse, self).__init__()
        self.obshii_massiv = parse_array('SELECT Name FROM Speller')
        self.id = int(id)
        if self.id > 3:
            self.id = 1

    @property
    def filterOrf(self):
        collection_key = {}
        for key, _item in Slovar_Parse.items():
            _item = _item[0]
            string_find = str()
            if int(self.id) == 1:
                string_find = str(_item[4] + " " + _item[6] + " " + _item[7] + " " + _item[8]).replace("-",
                                                                                                       " ").replace(",",
                                                                                                                    " ").replace(
                    "{", " ").replace(">", " ").replace("<", " ").replace("}", " ").replace("/", " ").replace('"',
                                                                                                              "").replace(".", " ").split(
                    " ")
            elif int(self.id) == 2:
                string_find = str(_item[4]).replace("-", " ").replace(",", " ").replace("{", " ").replace(">",
                                                                                                          " ").replace(".", " ").replace(
                    "<", " ").replace("}", " ").replace("/", " ").replace('"', "").split(" ")
            elif int(self.id) == 3:
                string_find = str(_item[6] + " " + _item[7] + " " + _item[8]).replace("-", " ").replace(",",
                                                                                                        " ").replace(".", " ").replace(
                    "{", " ").replace(">", " ").replace("<", " ").replace("}", " ").replace("/", " ").replace('"',
                                                                                                              "").replace(".", " ").split(
                    " ")
            temp_array = []
            for val in string_find:
                if (val != 'None') and (val.find('id') == -1) and (not val.isdigit().__eq__(True)) and (val != ""):
                    val = str(val).strip()
                    # Бинарный поиск в массиве
                    center = len(self.obshii_massiv) // 2
                    index_one = int(0)
                    index_max = len(self.obshii_massiv) - 1
                    while self.obshii_massiv[center] != val and index_one <= index_max:
                        if val > self.obshii_massiv[center]:
                            index_one = center + 1
                        else:
                            index_max = center - 1
                        center = (index_one + index_max) // 2
                    if index_one > index_max:
                        temp_array.append(val)
            if len(temp_array) > 0:
                collection_key[key] = temp_array

        collection = {}
        for keys in collection_key.keys():
            items = Slovar_Parse[keys][0]
            collection[keys] = [items, collection_key[keys]]

        return collection


# --- Обработчик Flask ----------------------------------------------
class Mass(object):
    def __init__(self):
        super(Mass, self).__init__()

    @staticmethod
    def Array_Parse():
        return ParseJson().make_array()



@app.route("/")
def index():
    return render_template("index.html", slovar=Slovar_Parse, datetimenow=datetime.datetime.now().date(), flag=1)

@app.route('/signUpUser/<id>', methods=['POST'])
def signUpUser(id):
    parse_array(
        "INSERT INTO Speller (Name) VALUES ('{}')".format(id))
    return '<dialog open>в словарь добавлен {} </dialog>'.format(id)


@app.route("/proverka/<id>")
def proverka(id = None):
    id_get = id
    if re.method == "GET":
        collection = QeryStatsFileParse(id_get).filterOrf
        return render_template("proverka.html", slovar=collection, datetimenow=datetime.datetime.now().date(), flag=0, fl=id_get)

@app.route("/attribyte/<id>")
def attribyte(id=None):
    slovar = Slovar_Parse[id][0]
    return render_template('attribyte.html', slovar=slovar)


# -------------------------Старт программы-------------------------------------------
if __name__ == '__main__':
    Slovar_Parse = Mass.Array_Parse()
    app.run( debug=True)
